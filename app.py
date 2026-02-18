import os
import uuid
import subprocess
import shutil
import tempfile
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO
from flask import Flask, request, render_template, send_file, jsonify
from werkzeug.utils import secure_filename
import threading
import json
import re
import time

try:
    from pdf2docx import Converter
    HAVE_PDF2DOCX = True
except Exception:
    HAVE_PDF2DOCX = False

try:
    from pdfminer.high_level import extract_text
    HAVE_PDFMINER = True
except Exception:
    HAVE_PDFMINER = False

try:
    import pandas as pd
    HAVE_PANDAS = True
except Exception:
    HAVE_PANDAS = False

try:
    from fpdf import FPDF
    HAVE_FPDF = True
except Exception:
    HAVE_FPDF = False

try:
    from openpyxl import load_workbook, Workbook
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

try:
    from docx2pdf import convert as docx2pdf_convert
    HAVE_DOCX2PDF = True
except Exception:
    HAVE_DOCX2PDF = False

try:
    from PIL import Image
    HAVE_PIL = True
except Exception:
    HAVE_PIL = False

# -------------------------
# FFmpeg PATH (IMPORTANT)
# -------------------------
FFMPEG_PATH = r"C:\ffmpeg-8.0-full_build\bin\ffmpeg.exe"  

# -------------------------
# APP SETUP
# -------------------------
app = Flask(__name__)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
DOWNLOAD_FOLDER = os.path.join(BASE_DIR, 'downloads')
PROGRESS_FOLDER = os.path.join(BASE_DIR, "progress")

# Create folders
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
os.makedirs(PROGRESS_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xlsx', 'txt', 'pptx', 'ppt', 'csv', 'xls', 'mp4', 'mkv', 'mov', 'jpg', 'jpeg', 'png', 'zip'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER

# -------------------------
# HELPERS
# -------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def unique_filename(original):
    safe = secure_filename(original)
    base, ext = os.path.splitext(safe)
    return f"{base}_{uuid.uuid4().hex}{ext}"

def converted_filename(original, new_ext):
    base, _ = os.path.splitext(secure_filename(original))
    return f"{base}_converted{new_ext}"

def _secs_from_hms(h, m, s):
    return int(h) * 3600 + int(m) * 60 + float(s)

def ffmpeg_monitor(input_abs, output_abs, job_id, ffmpeg_args):
    """
    Runs ffmpeg via subprocess.Popen and monitors stderr to extract progress info.
    Writes updates to PROGRESS_FOLDER/{job_id}.json
    """
    progress_path = os.path.join(PROGRESS_FOLDER, f"{job_id}.json")
    with open(progress_path, "w", encoding="utf-8") as f:
        json.dump({"status": "starting", "percent": 0, "frame": 0, "fps": 0, "bitrate": "", "speed": "", "time": "00:00:00", "eta": None}, f)

    probe_cmd = [FFMPEG_PATH, "-i", input_abs]
    probe = subprocess.run(probe_cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True).stderr
    duration_match = re.search(r"Duration:\s*(\d+):(\d+):(\d+\.\d+)", probe)
    total_seconds = None
    if duration_match:
        total_seconds = _secs_from_hms(duration_match.group(1), duration_match.group(2), duration_match.group(3))

    proc = subprocess.Popen(ffmpeg_args, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True, bufsize=1)

    time_re = re.compile(r"time=(\d+):(\d+):(\d+\.\d+)")
    frame_re = re.compile(r"frame=\s*(\d+)")
    fps_re = re.compile(r"fps=\s*([\d\.]+)")
    bitrate_re = re.compile(r"bitrate=\s*([\d\.kKmMbB/]+)")
    speed_re = re.compile(r"speed=\s*([\d\.x]+)")

    try:
        while True:
            line = proc.stderr.readline()
            if not line:
                if proc.poll() is not None:
                    break
                time.sleep(0.1)
                continue

            tm = time_re.search(line)
            fr = frame_re.search(line)
            fps = fps_re.search(line)
            br = bitrate_re.search(line)
            sp = speed_re.search(line)

            current_seconds = None
            percent = 0.0
            time_str = None
            if tm:
                h, m, s = tm.groups()
                current_seconds = _secs_from_hms(h, m, s)
                time_str = f"{int(h):02d}:{int(m):02d}:{float(s):06.3f}"
                if total_seconds and total_seconds > 0:
                    percent = min(100.0, (current_seconds / total_seconds) * 100.0)

            status_obj = {
                "status": "running",
                "percent": round(percent, 2),
                "frame": int(fr.group(1)) if fr else None,
                "fps": float(fps.group(1)) if fps else None,
                "bitrate": br.group(1) if br else "",
                "speed": sp.group(1) if sp else "",
                "time": time_str if time_str else "",
                "eta_seconds": None
            }

            if percent > 0 and current_seconds:
                try:
                    remaining = (100.0 - percent) * (current_seconds / percent)
                    status_obj["eta_seconds"] = int(remaining)
                except Exception:
                    status_obj["eta_seconds"] = None

            with open(progress_path, "w", encoding="utf-8") as f:
                json.dump(status_obj, f)

    except Exception as e:
        with open(progress_path, "w", encoding="utf-8") as f:
            json.dump({"status": "error", "error": str(e)}, f)
    finally:
        proc.wait()
        if os.path.exists(output_abs):
            final_size = os.path.getsize(output_abs)
            done_obj = {"status": "done", "percent": 100.0, "frame": None, "fps": None, "bitrate": "", "speed": "", "time": "", "eta_seconds": 0, "size": final_size, "output": output_abs}
        else:
            done_obj = {"status": "failed", "error": "output file missing"}
        with open(progress_path, "w", encoding="utf-8") as f:
            json.dump(done_obj, f)

# -------------------------
# Compression helpers (shared utilities)
# -------------------------
def _recompress_image_file(path, image_max_width, image_quality):
    """
    Recompress a single image file (replace original).
    Returns True if replaced, False otherwise.
    """
    if not HAVE_PIL:
        return False
    try:
        img = Image.open(path)
    except Exception:
        return False

    try:
        w, h = img.size
        if w > image_max_width:
            new_h = int((image_max_width / w) * h)
            img = img.resize((image_max_width, new_h), Image.LANCZOS)

        fmt = (img.format or "").upper()
       
        if fmt == "PNG" and (img.mode in ("RGBA", "LA") or "transparency" in getattr(img, "info", {})):
            # try to optimize PNG (lossless)
            img.save(path, format="PNG", optimize=True)
        else:
            rgb = img.convert("RGB")
            rgb.save(path, format="JPEG", quality=int(image_quality), optimize=True)
        return True
    except Exception:
        return False

def _strip_docprops_and_rezip(src_dir, out_path):
    """
    Create zip from src_dir while skipping common metadata files (docProps).
    """
    with ZipFile(out_path, 'w', ZIP_DEFLATED) as zout:
        for folder, _, files in os.walk(src_dir):
            for fname in files:
                absf = os.path.join(folder, fname)
                relf = os.path.relpath(absf, src_dir)
                
                if relf.startswith("docProps/"):
                    continue
                
                zout.write(absf, relf)
    return out_path

# -------------------------
# Format-specific compressors
# -------------------------
def compress_docx_file(input_path, output_path, image_max_width=1600, image_quality=70, remove_core_props=True):
    """
    Smart DOCX compression:
    - Unzip docx
    - Downscale images in word/media
    - Remove docProps if remove_core_props True
    - Rezip to output_path
    """
    if not HAVE_PIL:
        raise RuntimeError("Pillow is required for DOCX compression")

    tmpdir = tempfile.mkdtemp(prefix="docx_")
    try:
        with ZipFile(input_path, 'r') as zin:
            zin.extractall(tmpdir)

        media_dir = os.path.join(tmpdir, "word", "media")
        if os.path.isdir(media_dir):
            for fname in os.listdir(media_dir):
                full = os.path.join(media_dir, fname)
                if os.path.isfile(full):
                    _recompress_image_file(full, image_max_width, image_quality)

        # Optionally remove core properties files to strip metadata
        if remove_core_props:
            prop_paths = ["docProps/core.xml", "docProps/app.xml"]
            for p in prop_paths:
                fp = os.path.join(tmpdir, p)
                if os.path.exists(fp):
                    try:
                        os.remove(fp)
                    except Exception:
                        pass

        # Rezip while skipping docProps
        _strip_docprops_and_rezip(tmpdir, output_path)
        return output_path
    finally:
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass

def compress_pptx_file(input_path, output_path, image_max_width=1600, image_quality=70, remove_thumbnails=True, remove_core_props=True):
    """
    Smart PPTX compression:
    - Unzip pptx
    - Downscale images in ppt/media
    - Remove slide thumbnails and docProps
    - Rezip
    """
    if not HAVE_PIL:
        raise RuntimeError("Pillow is required for PPTX compression")

    tmpdir = tempfile.mkdtemp(prefix="pptx_")
    try:
        with ZipFile(input_path, 'r') as zin:
            zin.extractall(tmpdir)

        media_dir = os.path.join(tmpdir, "ppt", "media")
        if os.path.isdir(media_dir):
            for fname in os.listdir(media_dir):
                full = os.path.join(media_dir, fname)
                if os.path.isfile(full):
                    _recompress_image_file(full, image_max_width, image_quality)

        # Remove thumbnails if exist (commonly in ppt/ or thumbnails/)
        thumb_paths = [
            os.path.join(tmpdir, "docProps", "thumbnail.jpeg"),
            os.path.join(tmpdir, "thumbnail.jpeg"),
            os.path.join(tmpdir, "ppt", "thumbnails")
        ]
        for p in thumb_paths:
            if os.path.isdir(p):
                try:
                    shutil.rmtree(p)
                except Exception:
                    pass
            else:
                if os.path.exists(p):
                    try:
                        os.remove(p)
                    except Exception:
                        pass

        # Remove docProps
        if remove_core_props:
            prop_paths = ["docProps/core.xml", "docProps/app.xml"]
            for p in prop_paths:
                fp = os.path.join(tmpdir, p)
                if os.path.exists(fp):
                    try:
                        os.remove(fp)
                    except Exception:
                        pass

        _strip_docprops_and_rezip(tmpdir, output_path)
        return output_path
    finally:
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass

def compress_xlsx_file(input_path, output_path, image_max_width=1600, image_quality=70,
                       flatten_formulas=True, remove_core_props=True):

    if not HAVE_PIL:
        raise RuntimeError("Pillow is required for XLSX compression")

    tmpdir = tempfile.mkdtemp(prefix="xlsx_")
    try:
        working_input = input_path

        # 1) Flatten formulas only if requested
        if flatten_formulas:
            try:
                wb_vals = load_workbook(input_path, data_only=True)
                flat_path = os.path.join(tmpdir, "flattened.xlsx")
                wb_vals.save(flat_path)
                working_input = flat_path
            except Exception:
                working_input = input_path

        # 2) Unzip XLSX
        unzip_dir = os.path.join(tmpdir, "unzipped")
        os.makedirs(unzip_dir, exist_ok=True)

        with ZipFile(working_input, 'r') as zin:
            zin.extractall(unzip_dir)

        # 3) Aggressively recompress images in xl/media
        media_dir = os.path.join(unzip_dir, "xl", "media")
        if os.path.isdir(media_dir):
            for fname in os.listdir(media_dir):
                fpath = os.path.join(media_dir, fname)
                if not os.path.isfile(fpath):
                    continue

                try:
                    img = Image.open(fpath)

                    # Resize if too large
                    w, h = img.size
                    if w > image_max_width:
                        nh = int((image_max_width / w) * h)
                        img = img.resize((image_max_width, nh), Image.LANCZOS)

                    # Convert everything to JPEG aggressively
                    rgb = img.convert("RGB")

                    buf = BytesIO()
                    rgb.save(buf, "JPEG", optimize=True, quality=image_quality)

                    with open(fpath, "wb") as f:
                        f.write(buf.getvalue())

                except Exception:
                    continue

        # 4) Remove metadata
        if remove_core_props:
            docprops_dir = os.path.join(unzip_dir, "docProps")
            if os.path.isdir(docprops_dir):
                shutil.rmtree(docprops_dir, ignore_errors=True)

        # 5) Rezip with maximum compression
        with ZipFile(output_path, "w", compression=ZIP_DEFLATED, compresslevel=9) as zout:
            for folder, _, files in os.walk(unzip_dir):
                for file in files:
                    full_path = os.path.join(folder, file)
                    rel = os.path.relpath(full_path, unzip_dir)
                    zout.write(full_path, rel)

        # 6) Safety fallback: ensure final is not larger
        if os.path.getsize(output_path) > os.path.getsize(input_path):
            with ZipFile(output_path, "w", compression=ZIP_DEFLATED, compresslevel=9) as zout:
                with ZipFile(input_path, "r") as zin:
                    for zf in zin.infolist():
                        zout.writestr(zf.filename, zin.read(zf.filename))

        return output_path

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

# -------------------------
# ROUTES
# -------------------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route("/compression_tools")
def compression_tools():
    return render_template("compression_tools.html")

@app.route("/conversion_tools")
def conversion_tools():
    return render_template("conversion_tools.html")

# -----------------------------------------
# PDF → WORD
# -----------------------------------------
@app.route('/pdf_to_word', methods=['GET', 'POST'])
def pdf_to_word():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or uploaded.filename == '':
            return render_template('pdf_to_word.html', error="No file selected")

        if uploaded.filename.rsplit('.', 1)[1].lower() != 'pdf':
            return render_template('pdf_to_word.html', error="Please upload a PDF")

        if not HAVE_PDF2DOCX:
            return render_template('pdf_to_word.html', error="pdf2docx not installed")

        try:
            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(DOWNLOAD_FOLDER, os.path.splitext(unique)[0] + ".docx")

            cv = Converter(in_path)
            cv.convert(out_path)
            cv.close()

            download_name = converted_filename(uploaded.filename, ".docx")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('pdf_to_word.html', error=str(e))

    return render_template('pdf_to_word.html')

# -----------------------------------------
# WORD → PDF
# -----------------------------------------
@app.route('/word_to_pdf', methods=['GET', 'POST'])
def word_to_pdf():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or uploaded.filename == '':
            return render_template('word_to_pdf.html', error="No file selected")

        if uploaded.filename.rsplit('.', 1)[1].lower() not in {'doc', 'docx'}:
            return render_template('word_to_pdf.html', error="Upload a Word file")

        if not HAVE_DOCX2PDF:
            return render_template('word_to_pdf.html', error="docx2pdf not available")

        try:
            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(DOWNLOAD_FOLDER, os.path.splitext(unique)[0] + ".pdf")

            docx2pdf_convert(in_path, out_path)

            download_name = converted_filename(uploaded.filename, ".pdf")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('word_to_pdf.html', error=str(e))

    return render_template('word_to_pdf.html')

# -----------------------------------------
# EXCEL → PDF
# -----------------------------------------
@app.route('/excel_to_pdf', methods=['GET', 'POST'])
def excel_to_pdf():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or uploaded.filename == '':
            return render_template('excel_to_pdf.html', error="No file selected")

        if uploaded.filename.rsplit('.', 1)[1].lower() not in {'xlsx', 'xls'}:
            return render_template('excel_to_pdf.html', error="Upload Excel file")

        if not HAVE_OPENPYXL:
            return render_template('excel_to_pdf.html', error="openpyxl missing")

        try:
            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(DOWNLOAD_FOLDER, os.path.splitext(unique)[0] + ".pdf")

            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.units import inch
            from reportlab.pdfgen import canvas

            workbook = load_workbook(in_path, data_only=True)
            sheet = workbook.active

            c = canvas.Canvas(out_path, pagesize=landscape(letter))

            max_row = sheet.max_row
            max_col = sheet.max_column

            width, height = landscape(letter)
            cell_w = (width - inch) / max_col if max_col else width
            cell_h = (height - inch) / max_row if max_row else height

            c.setFont("Helvetica", 8)

            for r in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    val = sheet.cell(r, col).value
                    txt = "" if val is None else str(val)
                    c.drawString((col - 1) * cell_w + 20, height - (r * cell_h), txt)

            c.save()

            download_name = converted_filename(uploaded.filename, ".pdf")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('excel_to_pdf.html', error=str(e))

    return render_template('excel_to_pdf.html')

# -----------------------------------------
# PDF → CSV
# -----------------------------------------
@app.route('/pdf_to_csv', methods=['GET', 'POST'])
def pdf_to_csv():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or uploaded.filename == '':
            return render_template('pdf_to_csv.html', error="No file selected")

        if uploaded.filename.rsplit('.', 1)[1].lower() != 'pdf':
            return render_template('pdf_to_csv.html', error="Upload a PDF")

        try:
            import pdfplumber

            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(DOWNLOAD_FOLDER, os.path.splitext(unique)[0] + ".csv")

            rows = []
            with pdfplumber.open(in_path) as pdf:
                for page in pdf.pages:
                    table = page.extract_table()
                    if table:
                        rows.extend(table)

            if not rows:
                return render_template('pdf_to_csv.html', error="No table detected in PDF")

            df = pd.DataFrame(rows[1:], columns=rows[0])
            df.to_csv(out_path, index=False)

            download_name = converted_filename(uploaded.filename, ".csv")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('pdf_to_csv.html', error=str(e))

    return render_template('pdf_to_csv.html')

# -----------------------------------------
# PDF → EXCEL
# -----------------------------------------
@app.route('/pdf_to_excel', methods=['GET', 'POST'])
def pdf_to_excel():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or uploaded.filename == '':
            return render_template('pdf_to_excel.html', error="No file selected")

        if uploaded.filename.rsplit('.', 1)[1].lower() != 'pdf':
            return render_template('pdf_to_excel.html', error="Upload a PDF")

        try:
            import pdfplumber

            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(DOWNLOAD_FOLDER, os.path.splitext(unique)[0] + ".xlsx")

            rows = []
            with pdfplumber.open(in_path) as pdf:
                for page in pdf.pages:
                    table = page.extract_table()
                    if table:
                        rows.extend(table)

            if not rows:
                return render_template('pdf_to_excel.html', error="No table detected in PDF")

            df = pd.DataFrame(rows[1:], columns=rows[0])
            df.to_excel(out_path, index=False)

            download_name = converted_filename(uploaded.filename, ".xlsx")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('pdf_to_excel.html', error=str(e))

    return render_template('pdf_to_excel.html')

# -----------------------------------------
# PDF → TEXT
# -----------------------------------------
@app.route('/pdf_to_txt', methods=['GET', 'POST'])
def pdf_to_txt():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or uploaded.filename == '':
            return render_template('pdf_to_txt.html', error="No file selected")

        if uploaded.filename.rsplit('.', 1)[1].lower() != 'pdf':
            return render_template('pdf_to_txt.html', error="Upload a PDF")

        if not HAVE_PDFMINER:
            return render_template('pdf_to_txt.html', error="pdfminer missing")

        try:
            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(DOWNLOAD_FOLDER, os.path.splitext(unique)[0] + ".txt")

            text = extract_text(in_path)
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(text)

            download_name = converted_filename(uploaded.filename, ".txt")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('pdf_to_txt.html', error=str(e))

    return render_template('pdf_to_txt.html')

# --------------------
# TXT -> PDF (using FPDF + NotoSans)
# --------------------
@app.route('/txt_to_pdf', methods=['GET', 'POST'])
def txt_to_pdf():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or uploaded.filename == '':
            return render_template('txt_to_pdf.html', error="No file selected")

        if uploaded.filename.rsplit('.', 1)[1].lower() != 'txt':
            return render_template('txt_to_pdf.html', error="Please upload a .txt file")

        try:
            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(
                DOWNLOAD_FOLDER,
                os.path.splitext(unique)[0] + "_converted.pdf"
            )

            from fpdf import FPDF

            pdf = FPDF()
            pdf.add_page()

            # Use NotoSans-Regular.ttf (must exist in project folder)
            font_path = os.path.join(BASE_DIR, "NotoSans-Regular.ttf")
            if os.path.exists(font_path):
                pdf.add_font("Noto", "", font_path, uni=True)
                pdf.set_font("Noto", size=11)
            else:
                pdf.set_font("Helvetica", size=11)

            with open(in_path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    pdf.multi_cell(0, 8, line.rstrip())

            pdf.output(out_path)

            download_name = converted_filename(uploaded.filename, ".pdf")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('txt_to_pdf.html', error=f"Conversion failed: {e}")

    return render_template('txt_to_pdf.html')

# --------------------
# IMAGE COMPRESSION
# --------------------
@app.route('/image_compression', methods=['GET', 'POST'])
def image_compression():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        quality = request.form.get('quality', type=int)

        if not uploaded or uploaded.filename == '':
            return render_template('image_compression.html', error="No file selected")

        ext = uploaded.filename.rsplit('.', 1)[1].lower()
        if ext not in {'jpg', 'jpeg', 'png'}:
            return render_template('image_compression.html', error="Please upload JPG, JPEG, or PNG")

        try:
            if not HAVE_PIL:
                return render_template('image_compression.html', error="Pillow not installed")

            unique = unique_filename(uploaded.filename)
            in_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded.save(in_path)

            out_path = os.path.join(
                DOWNLOAD_FOLDER,
                os.path.splitext(unique)[0] + "_converted.jpg"
            )

            img = Image.open(in_path).convert("RGB")
            img.save(out_path, "JPEG", optimize=True, quality=quality)

            original_size = os.path.getsize(in_path)
            compressed_size = os.path.getsize(out_path)

            while compressed_size > original_size and quality > 10:
                quality -= 5
                img.save(out_path, "JPEG", optimize=True, quality=quality)
                compressed_size = os.path.getsize(out_path)

            download_name = converted_filename(uploaded.filename, ".jpg")
            return send_file(out_path, as_attachment=True, download_name=download_name)

        except Exception as e:
            return render_template('image_compression.html', error=f"Compression failed: {e}")

    return render_template('image_compression.html')

# --------------------
# VIDEO COMPRESSION
# --------------------
@app.route("/video_compression", methods=["GET", "POST"])
def video_compression():
    if request.method == "POST":
        file = request.files.get("file")
        user_quality = int(request.form.get("quality", 70))

        if not file or file.filename == "":
            return render_template("video_compression.html", error="No file selected")

        input_name = unique_filename(file.filename)
        input_path = os.path.join(UPLOAD_FOLDER, input_name)
        file.save(input_path)

        base, ext = os.path.splitext(input_name)
        output_name = f"{base}_compressed.mp4"
        output_path = os.path.join(DOWNLOAD_FOLDER, output_name)

        input_abs = os.path.abspath(input_path)
        output_abs = os.path.abspath(output_path)

        probe_cmd = [FFMPEG_PATH, "-i", input_abs]
        probe = subprocess.run(probe_cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True).stderr

        match = re.search(r"bitrate:\s*(\d+)\s*kb/s", probe)
        original_bitrate = int(match.group(1)) if match else 2000

        target_bitrate = int(original_bitrate * (user_quality / 100))
        target_bitrate = max(target_bitrate, 300)

        ffmpeg_args = [
            FFMPEG_PATH,
            "-i", input_abs,
            "-b:v", f"{target_bitrate}k",
            "-b:a", "64k",
            "-preset", "medium",
            "-y",
            output_abs
        ]

        job_id = uuid.uuid4().hex
        thread = threading.Thread(target=ffmpeg_monitor, args=(input_abs, output_abs, job_id, ffmpeg_args), daemon=True)
        thread.start()

        return render_template("video_progress.html", job_id=job_id,
                               output_name=converted_filename(file.filename, ".mp4"))

    return render_template("video_compression.html")

@app.route("/video_progress")
def video_progress():
    job_id = request.args.get("job_id")
    if not job_id:
        return jsonify({"error": "job_id required"}), 400

    progress_path = os.path.join(PROGRESS_FOLDER, f"{job_id}.json")
    if not os.path.exists(progress_path):
        return jsonify({"status": "notfound"}), 404

    try:
        with open(progress_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if data.get("status") == "done" and data.get("output"):
            data["download_url"] = "/download_by_path?path=" + data["output"]
        return jsonify(data)
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500

from urllib.parse import unquote

@app.route("/download_by_path")
def download_by_path():
    raw = request.args.get("path")
    if not raw:
        return "missing path", 400

    path = os.path.abspath(unquote(raw))

    if not path.startswith(os.path.abspath(DOWNLOAD_FOLDER) + os.sep):
        return "invalid path", 400

    if not os.path.exists(path):
        return "not found", 404

    return send_file(path, as_attachment=True)

# ---------------------------------------------------
# COMPRESS PDF (keeps existing ghostscript-based behavior)
# ---------------------------------------------------
def compress_pdf_with_ghostscript(input_path, output_path, quality='ebook'):
    """
    Uses Ghostscript to compress PDF.
    quality: one of screen, ebook, printer, prepress
    Requires ghostscript (gs) available on PATH
    """
    gs_bin = shutil.which("gs") or shutil.which("gswin64c") or shutil.which("gswin32c")
    if not gs_bin:
        raise RuntimeError("Ghostscript not found on PATH (gs or gswin64c)")

    gs_cmd = [
        gs_bin,
        "-sDEVICE=pdfwrite",
        "-dCompatibilityLevel=1.4",
        "-dPDFSETTINGS=/" + quality,
        "-dNOPAUSE",
        "-dQUIET",
        "-dBATCH",
        f"-sOutputFile={output_path}",
        input_path
    ]
    proc = subprocess.run(gs_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if proc.returncode != 0:
        raise RuntimeError(f"Ghostscript failed: {proc.stderr.strip()}")
    return output_path

@app.route('/compress_pdf', methods=['GET', 'POST'])
def compress_pdf():
    if request.method == 'POST':
        file = request.files.get("file")
        level = request.form.get("level", "ebook")

        if not file or file.filename == "":
            return render_template("tool_page.html", title="Compress PDF",
                                   subtitle="Reduce PDF size",
                                   error="No file selected",
                                   accepted_formats="PDF")

        ext = file.filename.rsplit(".", 1)[1].lower()
        if ext != "pdf":
            return render_template("tool_page.html", title="Compress PDF",
                                   subtitle="Reduce PDF size",
                                   error="Upload a PDF",
                                   accepted_formats="PDF")

        unique = unique_filename(file.filename)
        input_path = os.path.join(UPLOAD_FOLDER, unique)
        file.save(input_path)

        out_name = unique.replace(".pdf", "_compressed.pdf")
        output_path = os.path.join(DOWNLOAD_FOLDER, out_name)

        try:
            compress_pdf_with_ghostscript(input_path, output_path, quality=level)
            return send_file(output_path, as_attachment=True,
                             download_name=converted_filename(file.filename, ".pdf"))
        except Exception as e:
            return render_template("tool_page.html", title="Compress PDF",
                                   subtitle="Reduce PDF size",
                                   error=f"Compression failed: {e}",
                                   accepted_formats="PDF")

    return render_template("tool_page.html", title="Compress PDF",
                           subtitle="Reduce PDF size",
                           accepted_formats="PDF")

# ---------------------------------------------------
#compression routes (Word, PPT, Excel)
# ---------------------------------------------------

#----- word -----
@app.route('/compress_word', methods=['GET', 'POST'])
def compress_word_route():
    if request.method == 'POST':
        file = request.files.get("file")
        quality = int(request.form.get("quality", 70))
        maxwidth = int(request.form.get("maxwidth", 1600))

        if not file or file.filename == "":
            return render_template("word_compression.html", error="No file selected")

        ext = file.filename.rsplit(".", 1)[1].lower()
        if ext != "docx":
            return render_template("word_compression.html", error="Upload a DOCX file")

        unique = unique_filename(file.filename)
        input_path = os.path.join(UPLOAD_FOLDER, unique)
        file.save(input_path)

        out_name = os.path.splitext(unique)[0] + "_compressed.docx"
        output_path = os.path.join(DOWNLOAD_FOLDER, out_name)

        try:
            compress_docx_file(input_path, output_path, image_max_width=maxwidth, image_quality=quality)
            return send_file(output_path, as_attachment=True,
                             download_name=converted_filename(file.filename, ".docx"))
        except Exception as e:
            return render_template("word_compression.html", error=f"Compression failed: {e}")

    return render_template("word_compression.html")


#------- ppt -------

@app.route('/compress_ppt', methods=['GET', 'POST'])
def compress_ppt_route():
    if request.method == 'POST':
        file = request.files.get("file")
        quality = int(request.form.get("quality", 70))
        maxwidth = int(request.form.get("maxwidth", 1600))

        if not file or file.filename == "":
            return render_template("ppt_compression.html", error="No file selected")

        ext = file.filename.rsplit(".", 1)[1].lower()
        if ext != "pptx":
            return render_template("ppt_compression.html", error="Upload a PPTX file")

        unique = unique_filename(file.filename)
        input_path = os.path.join(UPLOAD_FOLDER, unique)
        file.save(input_path)

        out_name = os.path.splitext(unique)[0] + "_compressed.pptx"
        output_path = os.path.join(DOWNLOAD_FOLDER, out_name)

        try:
            compress_pptx_file(input_path, output_path,
                               image_max_width=maxwidth,
                               image_quality=quality)
            return send_file(output_path, as_attachment=True,
                             download_name=converted_filename(file.filename, ".pptx"))
        except Exception as e:
            return render_template("ppt_compression.html", error=f"Compression failed: {e}")

    return render_template("ppt_compression.html")


#----- excel -----

@app.route('/compress_excel', methods=['GET', 'POST'])
def compress_excel_route():
    if request.method == 'POST':
        file = request.files.get("file")
        quality = int(request.form.get("quality", 70))
        maxwidth = int(request.form.get("maxwidth", 1600))
        flatten = request.form.get("flatten", "on") == "on"

        if not file or file.filename == "":
            return render_template("excel_compression.html", error="No file selected")

        ext = file.filename.rsplit(".", 1)[1].lower()
        if ext not in {"xlsx"}:
            return render_template("excel_compression.html", error="Upload an XLSX file")

        unique = unique_filename(file.filename)
        input_path = os.path.join(UPLOAD_FOLDER, unique)
        file.save(input_path)

        out_name = os.path.splitext(unique)[0] + "_compressed.xlsx"
        output_path = os.path.join(DOWNLOAD_FOLDER, out_name)

        try:
            compress_xlsx_file(input_path, output_path,
                               image_max_width=maxwidth,
                               image_quality=quality,
                               flatten_formulas=flatten)
            return send_file(output_path, as_attachment=True,
                             download_name=converted_filename(file.filename, ".xlsx"))
        except Exception as e:
            return render_template("excel_compression.html", error=f"Compression failed: {e}")

    return render_template("excel_compression.html")

# ---------------------------------------------------
# CREATE ZIP (keeps existing behavior)
# ---------------------------------------------------
@app.route('/create_zip', methods=['GET', 'POST'])
def create_zip():
    if request.method == 'POST':
        files = request.files.getlist("file")

        if not files:
            return render_template("tool_page.html", title="Create ZIP",
                                   subtitle="Bundle multiple files",
                                   error="No files selected",
                                   accepted_formats="Any")

        zip_name = f"bundle_{uuid.uuid4().hex}.zip"
        zip_path = os.path.join(DOWNLOAD_FOLDER, zip_name)

        with ZipFile(zip_path, "w", ZIP_DEFLATED) as z:
            for f in files:
                if not f or f.filename == "":
                    continue
                unique = unique_filename(f.filename)
                saved_path = os.path.join(UPLOAD_FOLDER, unique)
                f.save(saved_path)
                z.write(saved_path, arcname=secure_filename(f.filename))

        return send_file(zip_path, as_attachment=True, download_name="bundle.zip")

    return render_template("tool_page.html", title="Create ZIP",
                           subtitle="Bundle multiple files",
                           accepted_formats="Any")

# -------------------------
# STATUS
# -------------------------
@app.route('/status')
def status():
    return jsonify({
        "pdf2docx": HAVE_PDF2DOCX,
        "pdfminer": HAVE_PDFMINER,
        "pandas": HAVE_PANDAS,
        "fpdf": HAVE_FPDF,
        "openpyxl": HAVE_OPENPYXL,
        "docx2pdf": HAVE_DOCX2PDF,
        "ffmpeg_path": os.path.exists(FFMPEG_PATH),
        "pillow": HAVE_PIL
    })

# -------------------------
# RUN SERVER
# -------------------------
if __name__ == '__main__':
    app.run(debug=True, port=5000)
